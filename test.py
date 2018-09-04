import openpyxl as xl

# -*- coding:utf-8 -*-
from itertools import product
import types
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import worksheet
from openpyxl.utils import range_boundaries
from itertools import product
#import regex as re
def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)
        print('testx')

        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
    

def writeborder( sheet, r1, c1,r2,c2 ):
#==============================================================================
#     header_fill = PatternFill( start_color='FFFFFF', end_color='FFFFFF', fill_type='solid' )
#     header_font = Font( size=11, bold=True )
      #header_align = Alignment( vertical='center', horizontal='center' )
      header_side = Side( border_style='thin', color='000000' )
      header_borderL = Border( left=header_side, right=None, top=header_side, bottom=header_side )
      header_borderR = Border( left=None, right=header_side, top=header_side, bottom=header_side )
      header_borderC = Border( left=None, right=None, top=header_side, bottom=header_side )
      for r in range(r1+1,r2):
          for c in range(c1+1,c1):
              sheet.cell( column=c, row=r ).border = header_borderC
    
     
#     sheet.cell( column=c, row=r ).fill = header_fill
#     sheet.cell( column=c, row=r ).font = header_font
#     sheet.cell( column=c, row=r ).alignment = header_align
      #sheet.cell( column=c, row=r ).border = header_border


#patch_worksheet()
wb = xl.load_workbook('test.xlsx')
patch_worksheet()
wb.save('testout.xlsx')
print('comple')
