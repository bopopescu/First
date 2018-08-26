# -*- coding: utf-8 -*-
"""
Created on Sat Jun 23 09:55:11 2018

@author: Administrator
"""
import numpy as np
import matplotlib.pyplot as plt
import math
import sympy as sp 
import scipy.optimize as op 
import gs
#==============================================================================
# import pyomo.environ as pe
#==============================================================================
from sympy.parsing.sympy_parser import parse_expr
from sympy import plot_implicit
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from PyQt5.QtWidgets import QTableWidgetItem as Qitem
from PyQt5.QtGui import QColor,QBrush

def writeNewValueGUI(tableWidget,i,j,value):
    newItem = Qitem(str(value))
    newItem.setBackground(QColor(0,255,0)) #RGB ,yellow
    tableWidget.setItem(i,j,newItem)

def setColorGUI(tableWidget,i,j):
    newItem = Qitem()
    newItem.setBackground(QColor(0,255,0)) #RGB ,yellow
    tableWidget.setItem(i,j,newItem)
    
def wipAngle(xs2,xs1,KBEW): # xs2-xs1 according to KBEW
        if KBEW =='+x':
            w2opt = xs2 - xs1
        elif gs.KBEW =='-x':
            w2opt =  360-xs2- xs1
        else:
            print('pls define KBEW')
        return w2opt
def point(A,Ap,Distance):# from A,Ap,D to B
    AAP= (Ap-A)
    if AAP[2]>0:
        Direction = (AAP)/np.linalg.norm(AAP)
    else:
        Direction = -(AAP)/np.linalg.norm(AAP)  
    B = A+ Direction*Distance
    return B
def Tolerance(dat,xm1,xm2):# dat i"BC", "ED", 'Delta','CD',"F_X", "F_Y", "F_Z","Fp_X", "Fp_Y", "Fp_Z",'FE',"A_X", "A_Y", "A_Z","Ap_X", "Ap_Y", "Ap_Z",'Distance',
    BC = dat[0]
    ED = dat[1]
    Delta = dat[2]
    CD = dat[3]

    F = np.array([dat[4],dat[5],dat[6]])
    Fp = np.array([dat[7],dat[8],dat[9]])
    FE = dat[10]
    E = point(F,Fp,FE)
    
    An = np.array([dat[11],dat[12],dat[13]])
    Ap = np.array([dat[14],dat[15],dat[16]])
    Distance = dat[17]
    Bn = point(An,Ap,Distance)
    
    M1Out = Output(BC,CD,ED,xm1,An,Bn,E,F) #degree
    beta0 = M1Out[1]
    NYS_Tp = M1Out[2]
# =============================================================================
    #--------second point
    M2Out = Output(BC,CD,ED,xm2,An,Bn,E,F)
    beta02 = M2Out[1]
    NYS_Tsp = M2Out[2]
# =============================================================================
#     B2=B2[0]
#     E2=E2[0]
#     B=B[0]
#     E=E[0]
# =============================================================================
    w2p=abs(beta0-beta02)

    List2=[w2p,NYS_Tp]

    return List2
 
def updateTolerance(Base,w2ErrorList,w2List,kpiList,i,t,Target,pos,neg,xm1,xm2):
    
    # Base,w2ErrorList,w2List,kpiList is array-like, will update them after running
    NumTolerance = len(Base) # tolerance parameter number
    zero = np.zeros(NumTolerance)
    Positive=zero.copy()
    Negative=zero.copy()
    Positive[i] = pos
    Negative[i] = neg
    ErrorPositive= Base + Positive
    ErrorNegative = Base - Negative

    w2p=Tolerance(ErrorPositive,xm1,xm2)[t]
    w2n=Tolerance(ErrorNegative,xm1,xm2)[t]
    
    w2PError = (w2p-Target)
    w2NError = (w2n-Target)
    
    if (abs(w2PError)>=abs(w2NError)):
        w2ErrorList.append(w2PError)
        w2List.append(w2p)
        kpi=ErrorPositive[i]
    else:
        w2ErrorList.append(w2NError)
        w2List.append(w2n)
        kpi=ErrorNegative[i]
    kpiList.append(kpi)
    Base[i]=kpi
    return



def OutputSymbol(A,B,E,F):
    BC,xLink,xCrank,xm, xs = sp.symbols('BC xLink xCrank xm xs',real=True) # link, outcrank, master angle, slave angle
    Bn=B.copy()
    An=A.copy()
    En=E.copy()
    Fn=F.copy()
    ab=Bn-An
    fe=En-Fn
    z0=ab/np.linalg.norm(-ab)
    z1=fe/np.linalg.norm(-fe)
    z0=allign(z0)
    z1=allign(z1)
    z0 = -z0 if (z0[2]>0) else z0
    x0= np.array([1,0,-z0[0]/z0[2]]) 
    x0=x0/np.linalg.norm(x0)
    y0=np.cross(z0,x0)
    y0=y0/np.linalg.norm(y0)
    C0=np.array([x0,y0,z0])
    p0r =np.array([BC*sp.cos(xm),BC*sp.sin(xm),0])
    p0r.shape=(1,3)
    Bn.shape=(1,3)
    ab.shape=(1,3)
    z1 = -z1 if (z1[2]>0) else z1
    x1= np.array([1,0,-z1[0]/z1[2]])
    x1=x1/np.linalg.norm(x1)
    y1=np.cross(z1,x1)
    y1=y1/np.linalg.norm(y1)
    C1=np.array([x1,y1,z1])
    p1r =np.array([xCrank*sp.cos(xs),xCrank*sp.sin(xs),0])
    p1r.shape=(1,3)
    En.shape=(1,3)
    
    CT=Bn.T+np.dot(C0.T,p0r.T)
    DT=En.T+np.dot(C1.T,p1r.T)
    
    #equation
    #1 BD vertical to AB
    D=DT.T
    C=CT.T
    cdt=DT-CT
    cd = cdt.T
    
    bd=D-B
    abt=ab.T
    
    # NYK_T

    DProject = project(D[0],ab[0],Bn[0]) # projcet to actua axis
    bdProject= DProject-Bn[0]
    bc=C[0]-Bn[0]
    
    CProject = project(C[0],fe,En[0])
    cdActuate = DProject-C[0]
    cdOut = D[0]-CProject
    de=En[0]-D[0]
    link=sum(cdt[i]**2 for i in range(3))
    equal1 = (link-xLink**2)/(xLink**2)
    equal2 =  (angle2(-bc,cdActuate))# bd  parallel to bc
    NYK_T = angle2(-bc,cdActuate)
    NYS_T = angle2(de,-cdOut)

    
    return [equal1,equal2,NYK_T,NYS_T,-cdOut[0],-cdOut[1],-cdOut[2]]

def angleDiff(angle1,angle2):
    angle =abs(angle1-angle2)
    if angle>360:
        print('angle difference is larger than 360')
    angle = angle if (angle <180) else (360-angle)
    return angle
def getMax(df,i):
    row=df.iloc[:,i].idxmax()
    a=df.iloc[row,0] 
    b=df.iloc[row,i]
    return[a,b]
def getMin(df,i):
    row=df.iloc[:,i].idxmin()
    a=df.iloc[row,0]
    b=df.iloc[row,i]
    return[a,b]



def getDegree(cxs,sxs):
    error=abs(cxs**2+sxs**2-1)
    if error>1e-8:
        print ('sin cos error:cxs=%f'%cxs,'sxs=%f'%sxs)
    if (sxs>=0):
        return degree(math.acos(cxs))
    else:
        return -degree(math.acos(cxs))
    
def allign(x):
    x.shape=(1,3)
    x=x[0]
    return x
def allignT(x):
    x.shape=(3,1)
    x=x[0]
    return x

def rad(x):
    return x*math.pi/180
def degree(x):
    return x*180/math.pi

def angle(a,b):
    a_b = sum(a[i]**2 for i in range (3))*sum(b[i]**2 for i in range (3))
    x=sum(a[i]*b[i] for i in range(3))/sp.sqrt(a_b)
    return sp.acos(x)
def angleMath(a,b):
    a_b = sum(a[i]**2 for i in range (3))*sum(b[i]**2 for i in range (3))
    x=sum(a[i]*b[i] for i in range(3))/math.sqrt(a_b)
    return math.acos(x)

def angle2(a,b):
    a_b = sum(a[i]**2 for i in range (3))*sum(b[i]**2 for i in range (3))
    x=sum(a[i]*b[i] for i in range(3))/sp.sqrt(a_b)
    return x

def project(A,vector,B): # A project to plane whose 法向量是vector ，并经过B点 
    u= sum((vector[i]*B[i]-vector[i]*A[i])for i in range(3))
    l =sum(vector[i]**2 for i in range(3))
    t= u/l

    projectx = A[0]+vector[0]*t
    projecty = A[1]+vector[1]*t
    projectz = A[2]+vector[2]*t

    return np.array([projectx,projecty,projectz])


def coordinate(A,AB,THETA_H,THETA_V):
    t1 = math.tan(THETA_H)
    t2= math.tan(THETA_V)
    direct = np.array([1,t1,t2])
    direct = direct/np.linalg.norm(direct)
    B =A-AB*direct
    if B.shape == 1:
        B=B[0]
    return B

#==============================================================================
def Output(BC,xLink,xOutCrank,alpha,A,B,E,F,KBEW='-x'): #xlink,xoutcrank, alpha,A,B,E,F array
#==============================================================================
    Bn = B.copy()
    An = A.copy()
    En = E.copy()
    Fn = F.copy()
    alpha =rad(alpha)
    Beta=sp.symbols('Beta',real=True)
    ab=Bn-An
    fe=En-Fn
    z0=ab/np.linalg.norm(-ab)
    z1=fe/np.linalg.norm(-fe)
    z0=allign(z0)
    z1=allign(z1)
    z0 = -z0 if (z0[2]>0) else z0
    x0= np.array([1,0,-z0[0]/z0[2]]) 
    x0=x0/np.linalg.norm(x0)
    y0=np.cross(z0,x0)
    y0=y0/np.linalg.norm(y0)
    C0=np.array([x0,y0,z0])
    p0r =np.array([BC*math.cos(alpha),BC*math.sin(alpha),0])
    p0r.shape=(1,3)
    Bn.shape=(1,3)
    ab.shape=(1,3)
    z1 = -z1 if (z1[2]>0) else z1
    x1= np.array([1,0,-z1[0]/z1[2]])
    x1=x1/np.linalg.norm(x1)
    y1=np.cross(z1,x1)
    y1=y1/np.linalg.norm(y1)
    C1=np.array([x1,y1,z1])
    p1r =np.array([xOutCrank*sp.cos(Beta),xOutCrank*sp.sin(Beta),0])
    p1r.shape=(1,3)
    En.shape=(1,3)
    
    # =============================================================================
    #     A0 = [[math.cos(alpha),-math.sin(alpha),0],[math.sin(alpha),math.cos(alpha),0],[0,0,1]] 
    #     A1 = [[sp.cos(Beta),-sp.sin(Beta),0],[sp.sin(Beta),sp.cos(Beta),0],[0,0,1]] 
    #     A0 = np.array(A0)
    #     A1=np.array(A1)
    # =============================================================================
    #CT=B.T+np.dot(np.dot(C0,A0),p0r.T)
    #DT=E.T+np.dot(np.dot(C1,A1),p1r.T) 
    
    # =============================================================================
    # C0New = np.row_stack((np.c_[C0.T,BCol],Nrow))
    # C1New = np.row_stack((np.c_[C1.T,ECol],Nrow))
    # =============================================================================

    CT=Bn.T+np.dot(C0.T,p0r.T)

    DT=En.T+np.dot(C1.T,p1r.T)
    #equation
    #1 BD 垂直AB
    D=DT.T
    C=CT.T
    cdt=DT-CT
    cd = cdt.T
    
    link=sum(cdt[i]**2 for i in range(3))
    Equal1=(link-xLink**2)/xLink**2
    F1=repr(Equal1[0]).replace('cos','sp.cos').replace('sin','sp.sin').replace('Beta','betaV[0]')
    # =============================================================================
    def Fbeta(betaV):
        return eval(F1)
    # =============================================================================
    if KBEW =='+x':

        beta= op.fsolve(Fbeta,[0])[0]
        
    elif KBEW =='-x':
        beta= op.fsolve(Fbeta,[math.pi])[0]
    else:
        print('KBEW is not a valid value')
    error =Fbeta([beta])
    if error>0.1:    
        print('cannot find beta at this point error=%f'%error+ '\CD=%f' %xLink+'\talfa=%f'  %(alpha*180/math.pi))
        print(F1)
    #回带
    # =============================================================================
    #     A1b = [[math.cos(beta),-math.sin(beta),0],[math.sin(beta),math.cos (beta),0],[0,0,1]] 
    #     A1b=np.array(A1b)
    # =============================================================================
    p1r =np.array([xOutCrank*sp.cos(beta),xOutCrank*sp.sin(beta),0])
    p1r.shape=(1,3)
    DTb=En.T+np.dot(C1.T,p1r.T)
    #equation
    #1 BD 垂直AB
    Db=DTb.T
    
    cdtb=DTb-CT
    cdb = Db-C
    bcb = (C-Bn)[0]
    deb = (En-Db)[0]
    fe.shape=(1,3)
    cdb.shape=(1,3)
    ab.shape=(1,3)
    Db.shape=(1,3)
    NYS_A = angle(-fe[0],cdb[0])-math.pi/2
    #NYK-A ab,cd
    NYK_A = angle(ab[0],cdb[0])-math.pi/2
    # NYK_T
        
    # =============================================================================
    DProject = project(Db[0],ab[0],Bn[0]) # projcet to actua axis
    # =============================================================================
    # =============================================================================
    CProject = project(C[0],fe[0],En[0])
    cdActuate = DProject-C[0]
    cdOut = Db[0]-CProject
    bdActuate=DProject-Bn[0]
    
    NYK_T = math.pi/2-angle(-bcb,cdActuate)
    NYS_T = math.pi/2-angle(deb,-cdOut)
    ang=[beta,NYS_T,NYS_A,NYK_T,NYK_A]
    [beta,NYS_T,NYS_A,NYK_T,NYK_A] =[x*180/math.pi for x in ang]
    alpha = degree(alpha)
# =============================================================================
#==============================================================================
    return [alpha,beta,NYS_T,NYS_A,NYK_T,NYK_A,C[0][0],C[0][1],C[0][2],Db[0][0],Db[0][1],Db[0][2]]
    #return [beta,C[0][0],C[0][1],C[0][2],Db[0][0],Db[0][1],Db[0][2]]
    #return Equal1
#==============================================================================
# =============================================================================


def Output2(A,B,E,F,BC,xLink,xOutCrank):
    alpha,Beta=sp.symbols('x[0] x[1]',real=True)
    ab=B-A
    fe=E-F
    z0=ab/np.linalg.norm(-ab)
    z1=fe/np.linalg.norm(-fe)
    z0=allign(z0)
    z1=allign(z1)
    z0 = -z0 if (z0[2]>0) else z0
    x0= np.array([1,0,-z0[0]/z0[2]])
    x0=x0/np.linalg.norm(x0)
    y0=np.cross(z0,x0)
    y0=y0/np.linalg.norm(y0)
    C0=np.array([x0,y0,z0])
    p0r =np.array([BC*sp.cos(alpha),BC*sp.sin(alpha),0])
    p0r.shape=(1,3)
    B.shape=(1,3)
    ab.shape=(1,3)
    z1 = -z1 if (z1[2]>0) else z1
    x1= np.array([1,0,-z1[0]/z1[2]])
    x1=x1/np.linalg.norm(x1)
    y1=np.cross(z1,x1)
    y1=y1/np.linalg.norm(y1)
    C1=np.array([x1,y1,z1])
    p1r =np.array([xOutCrank*sp.cos(Beta),xOutCrank*sp.sin(Beta),0])
    p1r.shape=(1,3)
    E.shape=(1,3)

    # =============================================================================
    #     A0 = [[math.cos(alpha),-math.sin(alpha),0],[math.sin(alpha),math.cos(alpha),0],[0,0,1]]
    #     A1 = [[sp.cos(Beta),-sp.sin(Beta),0],[sp.sin(Beta),sp.cos(Beta),0],[0,0,1]]
    #     A0 = np.array(A0)
    #     A1=np.array(A1)
    # =============================================================================
    #CT=B.T+np.dot(np.dot(C0,A0),p0r.T)
    #DT=E.T+np.dot(np.dot(C1,A1),p1r.T)

    # =============================================================================
    # C0New = np.row_stack((np.c_[C0.T,BCol],Nrow))
    # C1New = np.row_stack((np.c_[C1.T,ECol],Nrow))
    # =============================================================================
    CT=B.T+np.dot(C0.T,p0r.T)
    DT=E.T+np.dot(C1.T,p1r.T)

    #equation
    #1 BD 垂直AB
    D=DT.T
    C=CT.T
    cdt=DT-CT
    cd = cdt.T
    DProject = project(D[0],ab[0],B[0]) # projcet to actua axis
    bdProject= DProject-B[0]
    bc=C[0]-B[0]
    link=sum(cdt[i]**2 for i in range(3))
    Equal1=(link-xLink**2)/10000
    Equal2= 1000*(angle2(bc,bdProject)-1)# bd 和bc平行
    Equal22= 1000*(angle2(bc,bdProject)+1)# bd 和bc平行
    Fx1=repr(Equal1[0]).replace('cos','sp.cos').replace('sin','sp.sin')
    Fx2=repr(Equal2).replace('cos','sp.cos').replace('sin','sp.sin').replace('sqrt','sp.sqrt')
    Fx12=repr(Equal1[0]).replace('cos','sp.cos').replace('sin','sp.sin')
    Fx22=repr(Equal22).replace('cos','sp.cos').replace('sin','sp.sin').replace('sqrt','sp.sqrt')


    # =============================================================================
    # model.Con1=pe.Constraint(expr=eval(Fz1)==0)
    # model.Con2=pe.Constraint(expr=eval(Fz2)==0)

    # =============================================================================
    def Fbeta(x):
        return [eval(Fx1),eval(Fx2)]
    def Fbeta2(x):
        return [eval(Fx12),eval(Fx22)]
    # =============================================================================

    #result_obj = opt.solve(model)
    [alpha,beta]= op.fsolve(Fbeta,[1.2,0.2])
    [alpha2,beta2]= op.fsolve(Fbeta2,[alpha+1.57,beta])
    error=Fbeta([alpha,beta])
    error2=Fbeta2([alpha2,beta2])
    if (abs(sum(error)+sum(error2)))>0.1:
        print(error,error2)
    angle=[alpha,alpha2,beta,beta2]
    [alpha,alpha2,beta,beta2]=[degree(t) for t in angle]

    w=abs(beta2-beta)%360
    wa=abs(alpha2-alpha)
    w= (360-w) if (w>180) else w
    return [w,wa]
def LBC2(ED,Distance2,Delta2):    
    angle1=math.asin(ED*math.sin(Delta2)/Distance2)
    y=math.sin(angle1+Delta2)/math.sin(Delta2)*Distance2
    return y

def write( sheet, r, c, value ):
#==============================================================================
#     header_fill = PatternFill( start_color='FFFFFF', end_color='FFFFFF', fill_type='solid' )
#     header_font = Font( size=11, bold=True )
      header_align = Alignment( vertical='center', horizontal='center' )
      header_side = Side( border_style='thin', color='000000' )
      header_border = Border( left=header_side, right=header_side, top=header_side, bottom=header_side )
      sheet.cell( column=c, row=r, value=value )
#     sheet.cell( column=c, row=r ).fill = header_fill
#     sheet.cell( column=c, row=r ).font = header_font
#     sheet.cell( column=c, row=r ).alignment = header_align
      sheet.cell( column=c, row=r ).border = header_border
#==============================================================================
def writeResult( sheet, r, c, value ):
#==============================================================================
      header_fill = PatternFill( start_color='ccff00', end_color='ccff00', fill_type='solid' )
#     header_font = Font( size=11, bold=True )
      header_align = Alignment( vertical='center', horizontal='center' )
      header_side = Side( border_style='thin', color='000000' )
      header_border = Border( left=header_side, right=header_side, top=header_side, bottom=header_side )
      sheet.cell( column=c, row=r, value=value )
      sheet.cell( column=c, row=r ).fill = header_fill
#     sheet.cell( column=c, row=r ).font = header_font
#     sheet.cell( column=c, row=r ).alignment = header_align
      sheet.cell( column=c, row=r ).border = header_border

def read(sheet,r,c):
    value= sheet.cell(row=r,column=c).value
    if value=='/':
        value = 0
    return str(value)
#write(excelPath,Sheetname,startRow,startCol)

def update_lines(num, dataLines, lines, ax):
    for line, data in zip(lines, dataLines):
        # NOTE: there is no .set_data() for 3 dim data.
        Cx = data[0, num]
        Cy = data[1, num]
        Cz = data[2, num]
        Dx = data[3, num]
        Dy = data[4, num]
        Dz = data[5, num]
        temp = [[Cx, Dx], [Cy, Dy]]

        line.set_data(temp)
        line.set_3d_properties([Cz, Dz])
    return lines