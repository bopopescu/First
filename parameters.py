# -*- coding: utf-8 -*-
"""
Created on Sat Aug  4 21:56:52 2018
test
@author: Administrator
"""
import numpy as np
import os
import openpyxl as xl
import Func
from openpyxl.utils import get_column_letter
from openpyxl import worksheet
import json
import math
#from openpyxl.utils import range_boundarie

def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)
        #print('testx')

        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
class error():
    Offset = 0

class Inputs(error):
    errorP= error()
    errorN =error()
    Customer    = 'c'
    ProjectName = 'p'
    Department    = 'd'
    Drawing     = 'D'
    ValidDate   = 'V'
    Name        = 'N'
    Comment     = 'C'
    ParkWhere =  'P'
    DriveType =  'D'
    MechanicType = 'M' 

    isClock     = 'C1'
    startFrom   = 'CEPS'

    APS1   = 0.
    APS2   = 0.
    IPL    = 0.
    UWL    = 0.
    ALP    = 0.
    SP     = 0.
    KMP    = 0.

    Alfa   = 0.
    Offset = 0.
    A_X   =  0.
    A_Y   =  0.
    A_Z   =  0.
    Ap_X   = 0.
    Ap_Y   = 0.
    Ap_Z   = 0.
    errorP.Alfa   =  0.
    errorP.Offset =  0.
    errorP.A_X    =  0.
    errorP.A_Y    =  0.
    errorP.A_Z    =  0.
    errorP.Ap_X   =  0.
    errorP.Ap_Y   =  0.
    errorP.Ap_Z   =  0.
    errorN.Alfa   =  0.
    errorN.Offset =  0.
    errorN.A_X   =   0.
    errorN.A_Y   =   0.
    errorN.A_Z   =   0.
    errorN.Ap_X   =  0.
    errorN.Ap_Y   =  0.
    errorN.Ap_Z   =  0.
    Master =  'driver side'
    w2Target          = 0.
    w3Target          = 0.

    Distance   =  0.
    Delta = 0.
    FE    = 0.
    BC    = 0.
    CD    = 0.
    ED    = 0.
    F_X   = 0.
    F_Y   = 0.
    F_Z   = 0.
    Fp_X  = 0.
    Fp_Y  = 0.
    Fp_Z  = 0.
    errorP.Distance = 0.
    errorP.Delta    = 0.
    errorP.FE       = 0.
    errorP.BC       = 0.
    errorP.CD       = 0.
    errorP.ED       = 0.
    errorP.F_X      = 0.
    errorP.F_Y      = 0.
    errorP.F_Z      = 0.
    errorP.Fp_X     = 0.
    errorP.Fp_Y     = 0.
    errorP.Fp_Z     = 0.
    errorN.Distance = 0.
    errorN.Delta    = 0.
    errorN.FE       = 0.
    errorN.BC       = 0.
    errorN.CD       = 0.
    errorN.ED       = 0.
    errorN.F_X      = 0.
    errorN.F_Y      = 0.
    errorN.F_Z      = 0.
    errorN.Fp_X     = 0.
    errorN.Fp_Y     = 0.
    errorN.Fp_Z     = 0.
    KBEW = '+x'
    Ra_Rb = 0.
    w2opt = 0.

    Distance2   = 0.
    Delta2 =  0.
    FE2    =  0.
    BC2    =  0.
    CD2    =  0.
    ED2    =  0.
    F_X2   =  0.
    F_Y2   =  0.
    F_Z2   =  0.
    Fp_X2  =  0.
    Fp_Y2  =  0.
    Fp_Z2  =  0.
    errorP.Distance2 = 0.
    errorP.Delta2 = 0.
    errorP.FE2    = 0.
    errorP.BC2    = 0.
    errorP.CD2    = 0.
    errorP.ED2    = 0.
    errorP.F_X2   = 0.
    errorP.F_Y2   = 0.
    errorP.F_Z2   = 0.
    errorP.Fp_X2  = 0.
    errorP.Fp_Y2  = 0.
    errorP.Fp_Z2  = 0.
    errorN.Distance2 = 0.
    errorN.Delta2 = 0.
    errorN.FE2    = 0.
    errorN.BC2    = 0.
    errorN.CD2    = 0.
    errorN.ED2    = 0.
    errorN.F_X2   = 0.
    errorN.F_Y2   = 0.
    errorN.F_Z2   = 0.
    errorN.Fp_X2  = 0.
    errorN.Fp_Y2  = 0.
    errorN.Fp_Z2  = 0.
    KBEW2 = '+x'
    Ra_Rb2 = 0.
    w3opt = 0.

class Outputs():
    
    CordinateList =[]
    CordinateDetailList = []
    CD = 0.
    ED = 0.
    Ra_Rb = 0.
    BC2 = 0.
    CD2 = 0.
    ED2 = 0.
    Ra_Rb2 = 0.
    listOpt1 = []
    listOpt2 = []
    startFrom = ''
    w2Target = 0.
    w3Target = 0.
    parameterOpt = ''
    w2opt = 0.
    w3opt = 0.
    M1NYS_T = 0.
    M2NYS_T = 0.
    S1NYS_T = 0.
    S2NYS_T = 0.
    outKPIall= []
    outCordinateall = []
    maxArray =[]
    minArray=[]
    maxArray2 = []
    minArray2 = []
    step = 0.
    w2cal = 0
    w3cal = 0
    Park = 0 # abs park angle compared with 0 angle
    UWL1 = 0 # abs park angle compared with 0 angle
    UWL2 = 0 # abs park angle compared with 0 angle
        # Tolerance
    no = 0              # which crank to be calculated, 2 means master crank ,3 means slave crank(N2 column)
    Parameter=[]        # list to store the paramter to be used for tolerance check(parameter column)
    kpiList = []        # list to store the value of parameter(e.g BC) when reaching the largest deviation (Dimension column)
    w2List=[]           # list to store the value of target function(e.g w2) when reaching the largest deviation(Function value column)
    w2ErrorList = []    # list to store the deviation of target function
    num = 0. # number of points doing alphanumeric calculation
    index = 0.
    noCrank = ''

class GS(Inputs,Outputs):
    # Tolerance
    no = 0              # which crank to be calculated, 2 means master crank ,3 means slave crank(N2 column)
    Parameter=[]        # list to store the paramter to be used for tolerance check(parameter column)
    kpiList = []        # list to store the value of parameter(e.g BC) when reaching the largest deviation (Dimension column)
    w2List=[]           # list to store the value of target function(e.g w2) when reaching the largest deviation(Function value column)
    w2ErrorList = []    # list to store the deviation of target function
    #alphanumeric
    outKPIall= []       # dataframe to store the kpi info  after alphanumeric calculation        
    #cols =['alpha','beta','beta_s','beta_ss','NYS_T','NYS_A','NYK_T','NYK_A']+['beta2','beta_s2','beta_ss2','NYS_T2','NYS_A2','NYK_T2','NYK_A2']
    outCordinateall = []# dataframe to store the cordinate info after alphanumeric calculation
    # colsCordinate=['Cx','Cy','Cz','Dx','Dy','Dz']+['Cx2','Cy2','Cz2','Dx2','Dy2','Dz2']

    CordinateList =[]         #list to store main cordinate list(30 degree), format to be seen in designtable1 template
    CordinateDetailList = []  # list to store detail list ,format to be seen in designtable2 template

    w2cal = 0.
    w3cal = 0.


    listOpt1 = []
    listOpt2 = []
    A = np.ones(3)
    B = np.ones(3)
    E = np.ones(3)
    F = np.ones(3)
    Ap = np.ones(3)
    Fp2 = np.ones(3)
    Fp = np.ones(3)
    A2 = np.ones(3)
    B2 = np.ones(3)
    E2 = np.ones(3)
    F2 = np.ones(3)

    xm1 = 0.
    xm2 = 0.
    xm3 = 0.
    xm12 = 0.
    xm22 = 0.

    xs1 = 0.
    xs2 = 0.
    xs3 = 0.
    xs12 = 0.
    xs22 = 0.
    BC = 50.
    CD = 250.
    ED = 62.
    BC2 = 50.
    CD2 = 450.
    ED2 = 62.
    Distance = 10.
    Distance2 = 10.
    Delta = 0.
    excel_out = ''


    w2opt = 0.
    w3opt = 0.
    minArray = []
    maxArray = []
    minArray2 = []
    maxArray2 = []
    Alfa = 360.
    num = 0
    Delta = 0.
    Delta2 = 0.
    BC2 = 0.
    styleTime = ''
    w2Target = 0.
    w2cal = 0.
    w3Target = 0.
    w3cal = 0.
    w2actual = 0.
    w3actual = 0.
    outKPIall = 0.
    step = 0.
    Customer = ''
    ProjectName = ''
    Drawing = ''
    ValidDate = ''
    Name = ''
    Comment = ''
    ParkPosition = ''
    w2 = 0.
    w3 = 0.
    isClock = 1
    isStandard = ''
    Measured = ''
    zeroAngle = 0.
    APS1 = 0.
    APS2 = 0
    IPL = 0.
    UWL = 0.
    ALP = 0.
    SP = 0.
    KMP = 0.
    Offset = 0.
    ParkWhere = ''
    DriveType = ''
    MechanicType = ''
    Ra_Rb = 0.
    Ra_Rb2 = 0.
    KBEW = '-x'
    KBEW2 = '-x'
    startFrom = ''
    parameterOpt =''
    A_X = 0.
    A_Y = 0.
    A_Y = 0.
    A_Z = 0.
    B_X = 0.
    B_Y = 0.
    B_Z = 0.
    A_X2 = 0.
    A_Y2 = 0.
    A_Z2 = 0.
    B_X2 = 0.
    B_Y2 = 0.
    B_Z2 = 0.
    E_X = 0.
    E_Y = 0.
    E_Z = 0.
    F_X = 0.
    F_Y = 0.
    F_Z = 0.
    E_X2 = 0.
    E_Y2 = 0.
    E_Z2 = 0.
    F_X2 = 0.
    F_Y2 = 0.
    F_Z2 = 0.
    Ap_X = 0.
    Ap_Y = 0.
    Ap_Z = 0.
    Ap_X2 = 0.
    Ap_Y2 = 0.
    Ap_Z2 = 0.
    Fp_X = 0.
    Fp_Y = 0.
    Fp_Z = 0.
    Fp_X2 = 0.
    Fp_Y2 = 0.
    Fp_Z2 = 0.
    M1NYS_T= 0.
    M2NYS_T= 0.
    M1NYK_T= 0.
    M2NYK_T= 0.
    S1NYS_T= 0.
    S2NYS_T= 0.
    S1NYK_T= 0.
    S2NYK_T= 0.
    M1NYS_Tb= 0.
    M2NYS_Tb= 0.
    M1NYK_Tb= 0.
    M2NYK_Tb= 0.
    S1NYS_Tb= 0.
    S2NYS_Tb= 0.
    S1NYK_Tb= 0.
    S2NYK_Tb= 0.
    Alfab = 0.
    offsetAngleb = 0.
    EDb = 0.
    BCb = 0.
    ED2b = 0.
    BC2b = 0.
    CDb = 0.
    CD2b = 0.
    xs1b = 0.
    xs2b = 0.
    xm1b= 0.
    xm2b = 0.
    xs12b = 0.
    xs22b = 0.
    w2optb = 0.
    w3optb = 0.
    line_ani = ''
    CordinateList = []
    template_path = os.getcwd()+'/template.xlsx'  # read
    excel_out = ''
    excel_design1 = ''
    excel_design2 = ''
    listWrite2 = []
    cwd = os.getcwd()
    sheetNew = ''
    pause = False
    offsetAngle = 0.
    Master = 'Driver Side'
    
    DesignPath = {'Center': [cwd+'\LinkLines\Center\Center_LinkLines_DesignTable1.xlsx', cwd+'\LinkLines\Center\Center_LinkLines_DesignTable2.xlsx'],
                'Serial DS-PS': [cwd+'\LinkLines\Serial DS-PS\Serial_DS-PS_Linklines_DesignTable1.xlsx', cwd+'\LinkLines\Serial DS-PS\Serial_DS-PS_Linklines_DesignTable2.xlsx'],
                'Serial PS-DS': [cwd+'\LinkLines\Serial PS-DS\Serial_PS-DS_Linklines_DesignTable1.xlsx', cwd+'\LinkLines\Serial PS-DS\Serial_PS-DS_Linklines_DesignTable2.xlsx']}
    def loadtemplate(self):
        self.wb = xl.load_workbook(self.template_path)
        patch_worksheet()

        self.sheet1 = self.wb['Input']
        list1 =[[5,2,5],[5,9,12],[5,15,18],[12,5,8],[12,15,18],[14,7,8],[14,15,18],[16,6,7],[16,10,11],[16,17,18],[22,2,5],[22,11,14],[24,4,5],[24,9,10],
        [24,13,14],[26,4,5],[26,9,10],[26,13,14],[26,18,19],[29,2,4],[29,9,11],[29,15,17],[32,2,4],[32,9,11],[32,15,17],[38,2,4],[38,9,11],[38,15,17],
        [41,2,4],[41,9,11],[41,15,17],[44,2,4],[44,9,11],[44,15,17],[47,2,4],[47,9,11],[47,15,17],[49,4,5],[49,9,10],[49,17,19],[55,2,4],[55,9,11],[55,15,17],
        [58,2,4],[58,9,11],[58,15,17],[61,2,4],[61,9,11],[61,15,17],[64,2,4],[64,9,11],[64,15,17],[66,4,5],[66,9,10],[66,17,18],[19,14,20],[35,14,20],[21,7,8],[22,7,8],
        [21,16,17],[22,16,17]]

        list4 =[[5,2,8],[5,9,15]]
        list5 = [[14,2,8],[15,3,8],[16,3,8],[31,2,8],[32,3,8],[33,3,8]]
        
        for lists in [list1,list4,list5]:   
            for item in lists:
                x = item[0]
                item.insert(2,x)
        list1.extend([[70,2,71,5],[70,7,71,9],[70,11,71,13],[70,15,71,20]])
        self.sheet2 = self.wb['Optimization']
        self.sheet3 = self.wb['Input']
        self.sheet4 = self.wb['Output-Alphanumeric']
        self.sheet5 = self.wb['Output-Alphanumeric (2)']
        self.sheet6 = self.wb['Output-Tolerance calcuation']
        for item1 in list1:
            Func.writeborder(self.sheet1,item1)
        for item4 in list4:
            Func.writeborder(self.sheet4,item4)
        for item5 in list5:
            Func.writeborder(self.sheet5,item5)
            
    def loadDesignTable(self):        
        [self.excel_design1, self.excel_design2] = self.DesignPath[self.MechanicType]
        self.wb1 =  xl.load_workbook( self.excel_design1 )
        self.sheetDesign1 = self.wb1['Sheet1']
        self.wb2 =  xl.load_workbook( self.excel_design2 )
        self.sheetDesign2 = self.wb2['Sheet1']

 
    def calCordinate(self):
        degree=[self.w2,self.w3,self.Alfa,self.errorP.Alfa ,self.errorN.Alfa ,self.Offset,self.errorP.Offset ,self.errorN.Offset,self.APS1,
        self.APS2,self.IPL,self.UWL,self.ALP,self.SP ,self.KMP ]
        for i in range(len(degree)):
            if type(degree[i])==str:
                print('%s is  not a number ! will be set to 0'%degree[i])
                degree[i]=0
        radius = [x*math.pi/180 for x in degree]
        [self.w2,self.w3,self.Alfa,self.errorP.Alfa ,self.errorN.Alfa ,self.Offset,self.errorP.Offset ,self.errorN.Offset,self.APS1,
        self.APS2,self.IPL,self.UWL,self.ALP,self.SP ,self.KMP ]=radius
         
        self.A = np.array([self.A_X,self.A_Y,self.A_Z])
        self.Ap = np.array([self.Ap_X,self.Ap_Y,self.Ap_Z])
        self.F = np.array([self.F_X,self.F_Y,self.F_Z])
        self.Fp = np.array([self.Fp_X,self.Fp_Y,self.Fp_Z])
        self.F2 = np.array([self.F_X2,self.F_Y2,self.F_Z2])
        self.Fp2 = np.array([self.Fp_X2,self.Fp_Y2,self.Fp_Z2])
        self.B = Func.point(self.A,self.Ap,self.Distance)
        self.E = Func.point(self.F,self.Fp,self.FE)
        self.E2 = Func.point(self.F2,self.Fp2,self.FE2)
        
        if self.MechanicType =='Center':
            self.A2 = self.A.copy()
            self.Ap2 = self.Ap.copy()
            self.A_X2 = self.A_X
            self.A_Y2 = self.A_Y
            self.A_Z2 = self.A_Z
            self.Ap_X2 = self.Ap_X
            self.Ap_Y2 = self.Ap_Y
            self.Ap_Z2 = self.Ap_Z
            self.errorP.A_X2 = self.errorP.A_X
            self.errorP.A_Y2 = self.errorP.A_Y
            self.errorP.A_Z2 = self.errorP.A_Z
            self.errorP.Ap_X2 = self.errorP.Ap_X
            self.errorP.Ap_Y2 = self.errorP.Ap_Y
            self.errorP.Ap_Z2 = self.errorP.Ap_Z
            self.errorN.A_X2 = self.errorN.A_X
            self.errorN.A_Y2 = self.errorN.A_Y
            self.errorN.A_Z2 = self.errorN.A_Z
            self.errorN.Ap_X2 = self.errorN.Ap_X
            self.errorN.Ap_Y2 = self.errorN.Ap_Y
            self.errorN.Ap_Z2 = self.errorN.Ap_Z
            self.Distance2=(self.Distance+self.Distance2)
            self.B2 = Func.point(self.A2,self.Ap2,self.Distance2)
            self.parameterOpt = 'ED'
        else:
            self.A2 = self.E.copy()
            self.Ap2 = self.F.copy()
            self.A_X2 = self.E_X
            self.A_Y2 = self.E_Y
            self.A_Z2 = self.E_Z
            self.Ap_X2 = self.F_X
            self.Ap_Y2 = self.F_Y
            self.Ap_Z2 = self.F_Z
            self.errorP.A_X2 = self.errorP.Fp_X
            self.errorP.A_Y2 = self.errorP.Fp_Y
            self.errorP.A_Z2 = self.errorP.Fp_Z
            self.errorP.Ap_X2 = self.errorP.F_X
            self.errorP.Ap_Y2 = self.errorP.F_Y
            self.errorP.Ap_Z2 = self.errorP.F_Z
            self.errorN.A_X2 = self.errorN.Fp_X
            self.errorN.A_Y2 = self.errorN.Fp_Y
            self.errorN.A_Z2 = self.errorN.Fp_Z
            self.errorN.Ap_X2 = self.errorN.F_X
            self.errorN.Ap_Y2 = self.errorN.F_Y
            self.errorN.Ap_Z2 = self.errorN.F_Z
            self.B2 = Func.point(self.A2,self.Ap2,self.Distance2)
            self.parameterOpt = 'BC'
      
    def preTreat(self):
        
        self.calCordinate() # pre treatment for inputs
        self.loadDesignTable()
        self.excel_out=self.cwd+'\\output\\'+self.ProjectName+'Report_'+self.styleTime+'.xlsx'

    def printGS(self):
        print ('----------------------------master link info-----------------------------')
        print('CD=%.4f'% self.CDb +'\tED=%.4f'% self.EDb +'\tNYS_T1=%.4f'% self.M1NYS_Tb +'\tNYS_T2=%.4f'% self.M2NYS_Tb)
        print('NYK_T1={:.4f},  NYK_T2= {:.4f} ' .format(self.M1NYK_T,self.M2NYK_T))
        print('w2=%.4f'%self.w2optb+'\tw2Target=%.4f'%(180/math.pi*self.w2))
        if self.DriveType=='Reversing':
            print('offsetAngle = {:.4f}, Reversing angle={:.4f}'.format(self.offsetAngle,self.Alfa))

        print('After Rounding:CD=%.1f'% self.CD +'\tED=%.1f'% self.ED +'\tNYS_T1=%.4f'% self.M1NYS_T +'\tNYS_T2=%.4f'% self.M2NYS_T)
        print('NYK_T1={:.4f},  NYK_T2= {:.4f} ' .format(self.M1NYK_T,self.M2NYK_T))
        print('w2=%.4f'%self.w2opt+'\tw2Target=%.4f'%(180/math.pi*self.w2))
        if self.DriveType=='Reversing':
            print('offsetAngle = {:.4f}, Reversing angle={:.4f}'.format(self.offsetAngleb,self.Alfab))

        print ('----------------------------slave link info-----------------------------')
        print( 'BC2=%.4f'%self.BC2+'CD2=%.4f'% self.CD2 + '\tED2=%.4f'% self.ED2
            +'\nDelta=%.4f'%self.Delta2 +'\tD=%.4f'%self.Distance2)
        print('NYS_T1=%.4f'%self.S1NYS_T+'\tNYS_T2=%.4f'%self.S2NYS_T)
        print('NYK_T1=%.4f'%self.S1NYK_T+'\tNYK_T2=%.4f'%self.S2NYK_T)
        print('w3=%.4f'%self.w3opt+'\tw3Target=%.4f'%(180/math.pi*self.w3))
        print('After Rounding:'+'BC2=%.1f'%self.BC2+'\tCD2=%.1f'%self.CD2
            +'\nDelta=%.4f'%self.Delta2+'\tED2=%.1f'%self.ED2)

        print('NYS_T1=%.4f'%self.S1NYS_T+'\tNYS_T2=%.4f'%self.S2NYS_T)
        print('NYK_T1=%.4f'%self.S1NYK_T+'\tNYK_T2=%.4f'%self.S2NYK_T)
        print('w3=%.4f'%self.w3opt+'\tw3Target=%.4f'%(180/math.pi*self.w3))
            
    def writeGS(self):
            Func.write(self.sheet2, 2, 12, self.startFrom)
            Func.write(self.sheet2,8,6,self.w2Target)
            Func.write(self.sheet2,16,2,self.EDb)
            Func.write(self.sheet2,16,3,self.CDb)
            Func.write(self.sheet2,16,5,'%.4f'% self.xs1b)
            Func.write(self.sheet2,16,6,'%.4f'% self.xs2b)
            Func.write(self.sheet2,16,7,'%.4f'% self.w2optb)
            Func.write(self.sheet2,16,9,'%.4f'%self.M1NYS_Tb)
            Func.write(self.sheet2,16,10,'%.4f'%self.M2NYS_Tb)
            Func.write(self.sheet2,16,11,'%.4f'%(self.M1NYS_Tb+self.M2NYS_Tb))
            Func.write(self.sheet2,17,2,self.ED)
            Func.write(self.sheet2,17,3,self.CD)
            Func.write(self.sheet2,17,5,'%.4f'%self.xs1)
            Func.write(self.sheet2,17,6,'%.4f'%self.xs2)
            Func.write(self.sheet2,17,7,'%.4f'%self.w2opt)
            Func.write(self.sheet2,17,9,'%.4f'%self.M1NYS_T)
            Func.write(self.sheet2,17,10,'%.4f'%self.M2NYS_T)
            Func.write(self.sheet2,17,11,'%.4f'%(self.M1NYS_T+self.M2NYS_T))
            Func.write(self.sheet2,19,3,'%.4f'%self.w2opt)
            Func.write(self.sheet2,19,9,'%.4f'%self.M1NYS_T)
            Func.write(self.sheet2,19,10,'%.4f'%self.M2NYS_T)
            
            if self.MechanicType =='Center':
                Func.write(self.sheet2,37,2,self.ED2b)
            else:
                Func.write(self.sheet2,37,2,self.BC2b)
            Func.write(self.sheet2,29,6,self.w3Target)
            Func.write(self.sheet2,37,3,self.CD2b)
            Func.write(self.sheet2,37,5,self.xs12b)
            Func.write(self.sheet2,37,6,self.xs22b)
            Func.write(self.sheet2,37,7,self.w3optb)
            Func.write(self.sheet2,37,9,self.S1NYS_Tb)
            Func.write(self.sheet2,37,10,self.S2NYS_Tb)
            Func.write(self.sheet2,37,11,(self.S1NYS_Tb+self.S2NYS_Tb))

            if self.MechanicType =='Center':
                Func.write(self.sheet2,38,2,self.ED2)
                Func.write(self.sheet2, 31, 6, 'ED')

            else:
                Func.write(self.sheet2,38,2,self.BC2)
                Func.write(self.sheet2, 31, 6, 'BC')

            Func.write(self.sheet2,38,3,self.CD2)
            Func.write(self.sheet2,38,5,'%.2f'% self.xs12)
            Func.write(self.sheet2,38,6,'%.2f'% self.xs22)
            Func.write(self.sheet2,38,7,'%.2f'%self.w3opt)
            Func.write(self.sheet2,38,9,'%.2f'%self.S1NYS_T)
            Func.write(self.sheet2,38,10,'%.2f'%self.S2NYS_T)
            Func.write(self.sheet2,38,11,'%.2f'%(self.S1NYS_T+self.S2NYS_T))
            Func.write(self.sheet2,40,3,'%.2f'%self.w3opt)
            Func.write(self.sheet2,40,9,'%.2f'%self.S1NYS_T)
            Func.write(self.sheet2,40,10,'%.2f'%self.S2NYS_T)
    
    def startCal(self):
        if self.startFrom == 'APS1':
            self.zeroAngle = self.xm1-self.APS1
            self.offsetAngle = self.offsetAngle-self.APS1
            self.Alfa = self.Alfa+self.APS1
        elif self.startFrom =='APS2':
            self.zeroAngle = self.xm1-self.APS2
            self.offsetAngle = self.offsetAngle-self.APS2
            self.Alfa = self.Alfa+self.APS2
        else:
            self.zeroAngle = self.xm1
            self.Alfa = self.Alfa
            self.offsetAngle = self.offsetAngle

  


    

   