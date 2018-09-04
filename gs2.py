# -*- coding: utf-8 -*-
"""
Created on Sat Aug  4 21:56:52 2018
test
@author: Administrator
"""
import numpy as np
import os
import openpyxl as xl
import Func
from openpyxl.utils import get_column_letter
from openpyxl import worksheet
import json
#from openpyxl.utils import range_boundarie
print('load gs')
def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)
        #print('testx')

        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
class error():
    Offset = 0

class Inputs(error):
    errorP= error()
    errorN =error()
    Customer    = 'c'
    ProjectName = 'p'
    Department    = 'd'
    Drawing     = 'D'
    ValidDate   = 'V'
    Name        = 'N'
    Comment     = 'C'
    ParkWhere =  'P'
    DriveType =  'D'
    MechanicType = 'M' 

    isClock     = 'C1'
    startFrom   = 'CEPS'

    APS1   = 0.
    APS2   = 0.
    IPL    = 0.
    UWL    = 0.
    ALP    = 0.
    SP     = 0.
    KMP    = 0.

    Alfa   = 0.
    Offset = 0.
    A_X   =  0.
    A_Y   =  0.
    A_Z   =  0.
    Ap_X   = 0.
    Ap_Y   = 0.
    Ap_Z   = 0.
    errorP.Alfa   =  0.
    errorP.Offset =  0.
    errorP.A_X    =  0.
    errorP.A_Y    =  0.
    errorP.A_Z    =  0.
    errorP.Ap_X   =  0.
    errorP.Ap_Y   =  0.
    errorP.Ap_Z   =  0.
    errorN.Alfa   =  0.
    errorN.Offset =  0.
    errorN.A_X   =   0.
    errorN.A_Y   =   0.
    errorN.A_Z   =   0.
    errorN.Ap_X   =  0.
    errorN.Ap_Y   =  0.
    errorN.Ap_Z   =  0.
    Master =  'driver side'
    w2          = 0.
    w3          = 0.

    Distance   =  0.
    Delta = 0.
    FE    = 0.
    BC    = 0.
    CD    = 0.
    ED    = 0.
    F_X   = 0.
    F_Y   = 0.
    F_Z   = 0.
    Fp_X  = 0.
    Fp_Y  = 0.
    Fp_Z  = 0.
    errorP.Distance = 0.
    errorP.Delta    = 0.
    errorP.FE       = 0.
    errorP.BC       = 0.
    errorP.CD       = 0.
    errorP.ED       = 0.
    errorP.F_X      = 0.
    errorP.F_Y      = 0.
    errorP.F_Z      = 0.
    errorP.Fp_X     = 0.
    errorP.Fp_Y     = 0.
    errorP.Fp_Z     = 0.
    errorN.Distance = 0.
    errorN.Delta    = 0.
    errorN.FE       = 0.
    errorN.BC       = 0.
    errorN.CD       = 0.
    errorN.ED       = 0.
    errorN.F_X      = 0.
    errorN.F_Y      = 0.
    errorN.F_Z      = 0.
    errorN.Fp_X     = 0.
    errorN.Fp_Y     = 0.
    errorN.Fp_Z     = 0.
    KBEW = '+x'
    Ra_Rb = 0.
    w2actual = 0.

    Distance2   = 0.
    Delta2 =  0.
    FE2    =  0.
    BC2    =  0.
    CD2    =  0.
    ED2    =  0.
    F_X2   =  0.
    F_Y2   =  0.
    F_Z2   =  0.
    Fp_X2  =  0.
    Fp_Y2  =  0.
    Fp_Z2  =  0.
    errorP.Distance2 = 0.
    errorP.Delta2 = 0.
    errorP.FE2    = 0.
    errorP.BC2    = 0.
    errorP.CD2    = 0.
    errorP.ED2    = 0.
    errorP.F_X2   = 0.
    errorP.F_Y2   = 0.
    errorP.F_Z2   = 0.
    errorP.Fp_X2  = 0.
    errorP.Fp_Y2  = 0.
    errorP.Fp_Z2  = 0.
    errorN.Distance2 = 0.
    errorN.Delta2 = 0.
    errorN.FE2    = 0.
    errorN.BC2    = 0.
    errorN.CD2    = 0.
    errorN.ED2    = 0.
    errorN.F_X2   = 0.
    errorN.F_Y2   = 0.
    errorN.F_Z2   = 0.
    errorN.Fp_X2  = 0.
    errorN.Fp_Y2  = 0.
    errorN.Fp_Z2  = 0.
    KBEW2 = '+x'
    Ra_Rb2 = 0.


class GS(Inputs):
    
    A = np.ones(3)
    B = np.ones(3)
    E = np.ones(3)
    F = np.ones(3)
    Ap = np.ones(3)
    Fp2 = np.ones(3)
    Fp = np.ones(3)
    A2 = np.ones(3)
    B2 = np.ones(3)
    E2 = np.ones(3)
    F2 = np.ones(3)

    xm1 = 0.
    xm2 = 0.
    xm3 = 0.
    xm12 = 0.
    xm22 = 0.

    xs1 = 0.
    xs2 = 0.
    xs3 = 0.
    xs12 = 0.
    xs22 = 0.
    BC = 50.
    CD = 250.
    ED = 62.
    BC2 = 50.
    CD2 = 450.
    ED2 = 62.
    Distance = 10.
    Distance2 = 10.
    Delta = 0.
    excel_out = ''
    w2Target = 0.
    w2cal = 0.
    w2opt = 0.
    w3Target = 0.
    w3cal = 0.
    w3opt = 0.
    minArray = []
    maxArray = []
    minArray2 = []
    maxArray2 = []

    Alfa = 360.
    num = 0
    Delta = 0.
    Delta2 = 0.
    BC2 = 0.
    styleTime = ''
    w2Target = 0.
    w2cal = 0.
    w3Target = 0.
    w3cal = 0.
    w2actual = 0.
    w3actual = 0.
    outKPIall = 0.
    step = 0.
    aa = 1.
    bb = 1.
    Customer = ''
    ProjectName = ''
    Drawing = ''
    ValidDate = ''
    Name = ''
    Comment = ''
    ParkPosition = ''
    w2 = 0.
    w3 = 0.
    isClock = 1
    isStandard = ''
    Measured = ''
    zeroAngle = 0.
    APS1 = 0.
    APS2 = 0
    IPL = 0.
    UWL = 0.
    ALP = 0.
    SP = 0.
    KMP = 0.
    Offset = 0.
    ParkWhere = ''
    DriveType = ''
    MechanicType = ''
    Ra_Rb = 0.
    Ra_Rb2 = 0.
    KBEW = '-x'
    KBEW2 = '-x'
    startFrom = ''
    A_X = 0.
    A_Y = 0.
    A_Y = 0.
    A_Z = 0.
    B_X = 0.
    B_Y = 0.
    B_Z = 0.
    A_X2 = 0.
    A_Y2 = 0.
    A_Z2 = 0.
    B_X2 = 0.
    B_Y2 = 0.
    B_Z2 = 0.
    E_X = 0.
    E_Y = 0.
    E_Z = 0.
    F_X = 0.
    F_Y = 0.
    F_Z = 0.
    E_X2 = 0.
    E_Y2 = 0.
    E_Z2 = 0.
    F_X2 = 0.
    F_Y2 = 0.
    F_Z2 = 0.
    Ap_X = 0.
    Ap_Y = 0.
    Ap_Z = 0.
    Ap_X2 = 0.
    Ap_Y2 = 0.
    Ap_Z2 = 0.
    Fp_X = 0.
    Fp_Y = 0.
    Fp_Z = 0.
    Fp_X2 = 0.
    Fp_Y2 = 0.
    Fp_Z2 = 0.

    line_ani = ''
    listWrite = []
    template_path = os.getcwd()+'/template.xlsx'  # read
    template_design1 = os.getcwd()+'/template_DesignTable1.xlsx'  # re.ad
    template_design2 = os.getcwd()+'/template_DesignTable2.xlsx'  # read
    template_design1_Series = os.getcwd()+'/template_DesignTable1_Series.xlsx'  # re.ad
    excel_out = ''
    excel_design1 = ''
    excel_design2 = ''
    listWrite2 = []
    wb = xl.load_workbook(template_path)
    patch_worksheet()

    sheet1 = wb['Input']
    list1 =[[5,2,5],[5,9,12],[5,15,18],[12,5,8],[12,15,18],[14,7,8],[14,15,18],[16,6,7],[16,10,11],[16,17,18],[22,2,5],[22,11,14],[24,4,5],[24,9,10],
    [24,13,14],[26,4,5],[26,9,10],[26,13,14],[26,18,19],[29,2,4],[29,9,11],[29,15,17],[32,2,4],[32,9,11],[32,15,17],[38,2,4],[38,9,11],[38,15,17],
    [41,2,4],[41,9,11],[41,15,17],[44,2,4],[44,9,11],[44,15,17],[47,2,4],[47,9,11],[47,15,17],[49,4,5],[49,9,10],[49,17,19],[55,2,4],[55,9,11],[55,15,17],
    [58,2,4],[58,9,11],[58,15,17],[61,2,4],[61,9,11],[61,15,17],[64,2,4],[64,9,11],[64,15,17],[66,4,5],[66,9,10],[66,17,18],[19,14,20],[35,14,20],[21,7,8],[22,7,8],
    [21,16,17],[22,16,17]]

    list4 =[[5,2,8],[5,9,15]]
    list5 = [[14,2,8],[15,3,8],[16,3,8],[31,2,8],[32,3,8],[33,3,8]]
    for lists in [list1,list4,list5]:   
        for item in lists:
            x = item[0]
            item.insert(2,x)
    list1.extend([[70,2,71,5],[70,7,71,9],[70,11,71,13],[70,15,71,20]])
    sheet2 = wb['Optimization']
    sheet3 = wb['Input']
    sheet4 = wb['Output-Alphanumeric']
    sheet5 = wb['Output-Alphanumeric (2)']
    sheet6 = wb['Output-Tolerance calcuation']
    for item1 in list1:
        Func.writeborder(sheet1,item1)
    for item4 in list4:
        Func.writeborder(sheet4,item4)
    for item5 in list5:
        Func.writeborder(sheet5,item5)
        

    wb1 = ''
    sheetDesign1 = ''
    wb2 = ''
    sheetDesign2 = ''
    wb1Series = ''
    sheetDesign1Series = ''
    ordinatePath = 'outputCordinate.xlsx'
    sheetNew = ''
    pause = False
    offsetAngle = 0.
    Master = 'Driver Side'
    cwd = os.getcwd()
    DesignPath = {'Center': [cwd+'\LinkLines\Center\Center_LinkLines_DesignTable1.xlsx', cwd+'\LinkLines\Center\Center_LinkLines_DesignTable2.xlsx'],
                'Serial DS-PS': [cwd+'\LinkLines\Serial DS-PS\Serial_DS-PS_Linklines_DesignTable1.xlsx', cwd+'\LinkLines\Serial DS-PS\Serial_DS-PS_Linklines_DesignTable2.xlsx'],
                'Serial PS-DS': [cwd+'\LinkLines\Serial PS-DS\Serial_PS-DS_Linklines_DesignTable1.xlsx', cwd+'\LinkLines\Serial PS-DS\Serial_PS-DS_Linklines_DesignTable2.xlsx']}
    def _calCordinate(self):
        A = np.array([self.A_X,self.A_Y,self.A_Z])
        Ap = np.array([self.Ap_X,self.Ap_Y,self.Ap_Z])
        F = np.array([self.F_X,self.F_Y,self.F_Z])
        Fp = np.array([self.Fp_X,self.Fp_Y,self.Fp_Z])
        F2 = np.array([self.F_X2,self.F_Y2,self.F_Z2])
        Fp2 = np.array([self.Fp_X2,self.Fp_Y2,self.Fp_Z2])
        B = Func.point(self.A,self.Ap,self.Distance)
        E = Func.point(self.F,self.Fp,self.FE)
        E2 = Func.point(self.F2,self.Fp2,self.FE2)
        
        if self.MechanicType =='Center':
            A2 = A.copy()
            Ap2 = Ap.copy()
            A_X2 = self.A_X
            A_Y2 = self.A_Y
            A_Z2 = self.A_Z
            Ap_X2 = self.Ap_X
            Ap_Y2 = self.Ap_Y
            Ap_Z2 = self.Ap_Z
            errorP.A_X2 = self.errorP.A_X
            errorP.A_Y2 = self.errorP.A_Y
            errorP.A_Z2 = self.errorP.A_Z
            errorP.Ap_X2 = self.errorP.Ap_X
            errorP.Ap_Y2 = self.errorP.Ap_Y
            errorP.Ap_Z2 = self.errorP.Ap_Z
            errorN.A_X2 = self.errorN.A_X
            errorN.A_Y2 = self.errorN.A_Y
            errorN.A_Z2 = self.errorN.A_Z
            errorN.Ap_X2 = self.errorN.Ap_X
            errorN.Ap_Y2 = self.errorN.Ap_Y
            errorN.Ap_Z2 = self.errorN.Ap_Z
            Distance2=(self.Distance+self.Distance2)
            B2 = Func.point(A2,Ap2,Distance2)
        else:
            A2 = E.copy()
            Ap2 = F.copy()
            A_X2 = self.E_X
            A_Y2 = self.E_Y
            A_Z2 = self.E_Z
            Ap_X2 = self.F_X
            Ap_Y2 = self.F_Y
            Ap_Z2 = self.F_Z
            errorP.A_X2 = self.errorP.Fp_X
            errorP.A_Y2 = self.errorP.Fp_Y
            errorP.A_Z2 = self.errorP.Fp_Z
            errorP.Ap_X2 = self.errorP.F_X
            errorP.Ap_Y2 = self.errorP.F_Y
            errorP.Ap_Z2 = self.errorP.F_Z
            errorN.A_X2 = self.errorN.Fp_X
            errorN.A_Y2 = self.errorN.Fp_Y
            errorN.A_Z2 = self.errorN.Fp_Z
            errorN.Ap_X2 = self.errorN.F_X
            errorN.Ap_Y2 = self.errorN.F_Y
            errorN.Ap_Z2 = self.errorN.F_Z
            B2 = Func.point(A2,Ap2,Distance2)
      

gs = GS()
selfs = Inputs()
jsonPath = 'input.json'
gs.ALP=30
def extractattr(c):
    name = dir(c)
    delname=[]
    for item in name:
        if item[0]=='_':
            delname.append(item)
    for item in delname:
        name.remove(item)
    return name

def writeInputjson(jsonPath,input,gs):
    nameInput = extractattr(input)
    nameError = extractattr(input.errorP)
    DictInput  ={}
    DictErrorP ={}
    DictErrorN ={}
    for var in nameError:
        DictErrorP[var]= getattr(gs.errorP, var)
        DictErrorN[var] = getattr(gs.errorN, var)
    for variable in nameInput:
        DictInput[variable]=getattr(gs,variable)
    DictInput['errorP'] = DictErrorP
    DictInput['errorN'] = DictErrorN
    jStr = json.dumps(DictInput)
    with open(jsonPath,'w') as f:
        f.write(jStr)

writeInputjson(jsonPath,inputs,gs)